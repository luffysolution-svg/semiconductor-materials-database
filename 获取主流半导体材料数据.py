"""
Materials Project API - 主流半导体材料数据获取
作者: Luffy.Solution
日期: 2025年10月6日
功能: 获取100+主流半导体材料的完整数据

Materials Project 数据使用 CC BY 4.0 许可。
数据来源: https://materialsproject.org/
"""

import json
import time
from datetime import datetime

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# API配置
try:
    from config import API_KEY, BASE_URL

    headers = {"X-API-KEY": API_KEY}
except ImportError:
    print("=" * 80)
    print("错误: 找不到 config.py 文件")
    print("=" * 80)
    print()
    print("请按照以下步骤配置:")
    print("1. 复制 config_example.py 为 config.py")
    print("2. 在 config.py 中填入您的 Materials Project API Key")
    print("   获取 API Key: https://materialsproject.org/api")
    print("3. 重新运行此脚本")
    print()
    print("=" * 80)
    exit(1)

print("=" * 80)
print("Materials Project API - 主流半导体材料数据获取系统")
print("=" * 80)
print()

# 定义主流半导体材料类别和搜索元素
SEMICONDUCTOR_CATEGORIES = {
    "金属硫化物": {
        "elements": [
            "Zn-S",
            "Cd-S",
            "Cu-S",
            "Mo-S",
            "W-S",
            "Sn-S",
            "Sb-S",
            "Bi-S",
            "In-S",
            "Ga-S",
        ],
        "description": "Metal Sulfides",
    },
    "金属氧化物": {
        "elements": [
            "Ti-O",
            "Zn-O",
            "Cu-O",
            "Fe-O",
            "Ni-O",
            "Co-O",
            "Sn-O",
            "Bi-O",
            "W-O",
            "V-O",
            "Ga-O",
            "In-O",
        ],
        "description": "Metal Oxides",
    },
    "金属硫氧化物": {
        "elements": ["Bi-S-O", "Cu-S-O", "Mo-S-O", "W-S-O"],
        "description": "Metal Oxysulfides",
    },
    "氮化物": {
        "elements": ["Ga-N", "In-N", "Al-N", "Ti-N", "C-N", "B-N", "Si-N"],
        "description": "Nitrides",
    },
    "碳化物": {
        "elements": ["Si-C", "Ti-C", "W-C", "Mo-C", "B-C"],
        "description": "Carbides",
    },
    "硒化物": {
        "elements": ["Cd-Se", "Zn-Se", "Cu-Se", "Mo-Se", "W-Se", "Bi-Se", "Sb-Se"],
        "description": "Selenides",
    },
    "碲化物": {
        "elements": ["Cd-Te", "Zn-Te", "Bi-Te", "Sb-Te", "Pb-Te"],
        "description": "Tellurides",
    },
    "卤化物": {
        "elements": ["Cu-I", "Ag-Br", "Pb-I", "Cs-Pb-I", "Cs-Pb-Br"],
        "description": "Halides (Perovskites)",
    },
    "磷化物": {
        "elements": ["Ga-P", "In-P", "Al-P", "Cd-P", "Zn-P"],
        "description": "Phosphides",
    },
    "砷化物": {"elements": ["Ga-As", "In-As", "Al-As"], "description": "Arsenides"},
}


def search_semiconductors_by_category(category_name, chemsys_list, limit_per_system=3):
    """
    按类别搜索半导体材料

    参数:
        category_name: 类别名称
        chemsys_list: 化学系统列表
        limit_per_system: 每个系统的材料数量限制
    """
    print(f"\n{'=' * 80}")
    print(f"正在搜索: {category_name}")
    print(f"{'=' * 80}")

    all_materials = []

    for chemsys in chemsys_list:
        print(f"\n  → 搜索化学系统: {chemsys}...", end=" ")

        url = f"{BASE_URL}/materials/summary/"

        params = {
            "chemsys": chemsys,
            "is_stable": True,  # 只要稳定相
            "band_gap_min": 0.1,  # 最小带隙
            "band_gap_max": 6.0,  # 最大带隙（排除绝缘体）
            "_fields": "material_id,formula_pretty,band_gap,is_gap_direct,energy_above_hull,"
            + "formation_energy_per_atom,density,volume,nsites,elements,nelements,"
            + "symmetry,efermi,is_metal,crystal_system,spacegroup_symbol",
            "_sort_fields": "energy_above_hull",  # 按稳定性排序
            "_limit": limit_per_system * 2,  # 多获取一些以备筛选
        }

        try:
            response = requests.get(url, headers=headers, params=params, timeout=30)

            if response.status_code == 200:
                data = response.json().get("data", [])

                # 筛选符合条件的材料
                filtered_materials = []
                for mat in data:
                    # 确保带隙在合理范围内
                    bg = mat.get("band_gap")
                    if bg and 0.1 <= bg <= 6.0 and not mat.get("is_metal", False):
                        mat["category"] = category_name
                        mat["chemsys"] = chemsys
                        filtered_materials.append(mat)

                        if len(filtered_materials) >= limit_per_system:
                            break

                all_materials.extend(filtered_materials)
                print(f"✓ 找到 {len(filtered_materials)} 个材料")

            else:
                print(f"✗ 请求失败 (状态码: {response.status_code})")

        except Exception as e:
            print(f"✗ 错误: {e}")

        # 速率限制
        time.sleep(0.5)

    print(f"\n{category_name} 共获取: {len(all_materials)} 个材料")
    return all_materials


def get_electronic_structure(material_id):
    """获取电子结构信息"""
    url = f"{BASE_URL}/materials/electronic_structure/"
    params = {
        "material_ids": material_id,
        "_fields": "material_id,band_gap,cbm,vbm,is_gap_direct,efermi,is_metal",
    }

    try:
        response = requests.get(url, headers=headers, params=params, timeout=20)
        if response.status_code == 200:
            data = response.json().get("data", [])
            return data[0] if data else {}
    except:
        pass
    return {}


def enrich_material_data(materials):
    """
    丰富材料数据，获取电子结构信息
    """
    print(f"\n{'=' * 80}")
    print("正在获取详细电子结构信息...")
    print(f"{'=' * 80}\n")

    enriched_materials = []

    for i, mat in enumerate(materials, 1):
        material_id = mat.get("material_id")
        print(
            f"[{i}/{len(materials)}] {mat.get('formula_pretty', 'Unknown')} ({material_id})...",
            end=" ",
        )

        # 获取电子结构
        elec_data = get_electronic_structure(material_id)

        # 合并数据
        enriched_mat = {**mat, **elec_data}
        enriched_materials.append(enriched_mat)

        print("✓")

        # 速率限制
        time.sleep(0.3)

        # 每10个显示进度
        if i % 10 == 0:
            print(f"   进度: {i}/{len(materials)} ({i / len(materials) * 100:.1f}%)")

    return enriched_materials


def create_dataframe(materials):
    """
    创建DataFrame并整理数据
    """
    print(f"\n{'=' * 80}")
    print("正在整理数据...")
    print(f"{'=' * 80}\n")

    df_data = []

    for mat in materials:
        row = {
            "分类": mat.get("category", "Unknown"),
            "化学系统": mat.get("chemsys", "Unknown"),
            "材料ID": mat.get("material_id", ""),
            "化学式": mat.get("formula_pretty", ""),
            "元素组成": ", ".join(mat.get("elements", [])),
            "元素数": mat.get("nelements", 0),
            "原子数": mat.get("nsites", 0),
            # 电子性质
            "带隙 (eV)": round(mat.get("band_gap", 0), 3)
            if mat.get("band_gap")
            else "N/A",
            "直接带隙": "是" if mat.get("is_gap_direct") else "否",
            "导带底 CBM (eV)": round(mat.get("cbm", 0), 3) if mat.get("cbm") else "N/A",
            "价带顶 VBM (eV)": round(mat.get("vbm", 0), 3) if mat.get("vbm") else "N/A",
            "费米能级 (eV)": round(mat.get("efermi", 0), 3)
            if mat.get("efermi")
            else "N/A",
            # 热力学性质
            "形成能 (eV/atom)": round(mat.get("formation_energy_per_atom", 0), 4)
            if mat.get("formation_energy_per_atom")
            else "N/A",
            "能量高于凸包 (eV/atom)": round(mat.get("energy_above_hull", 0), 4)
            if mat.get("energy_above_hull")
            else "N/A",
            # 结构性质
            "晶系": mat.get("crystal_system", "Unknown"),
            "空间群": mat.get("spacegroup_symbol", "Unknown"),
            "密度 (g/cm³)": round(mat.get("density", 0), 3)
            if mat.get("density")
            else "N/A",
            "体积 (Ų)": round(mat.get("volume", 0), 2) if mat.get("volume") else "N/A",
            # 应用潜力评估
            "光电应用潜力": assess_photovoltaic_potential(mat),
            "光催化应用潜力": assess_photocatalytic_potential(mat),
        }

        df_data.append(row)

    df = pd.DataFrame(df_data)

    # 按分类和带隙排序
    df = df.sort_values(by=["分类", "带隙 (eV)"], ascending=[True, True])
    df = df.reset_index(drop=True)

    return df


def assess_photovoltaic_potential(mat):
    """评估光伏应用潜力"""
    bg = mat.get("band_gap")
    energy_hull = mat.get("energy_above_hull", 1)

    if not bg or bg == "N/A":
        return "未知"

    if 1.1 <= bg <= 1.8 and energy_hull < 0.05:
        return "优秀"
    elif 0.8 <= bg <= 2.5 and energy_hull < 0.1:
        return "良好"
    elif 0.5 <= bg <= 3.0:
        return "一般"
    else:
        return "较低"


def assess_photocatalytic_potential(mat):
    """评估光催化应用潜力"""
    bg = mat.get("band_gap")
    cbm = mat.get("cbm")
    vbm = mat.get("vbm")

    if not bg or bg == "N/A":
        return "未知"

    # 水分解需要带隙 > 1.23 eV，且能带位置合适
    if bg and bg > 1.8 and bg < 3.5:
        if cbm and vbm:
            # 简化评估：CBM要足够负，VBM要足够正
            if cbm < 0 and vbm > -2.5:
                return "优秀"
            elif cbm < 0.5:
                return "良好"
        return "一般"
    elif bg and 1.5 <= bg <= 4.0:
        return "一般"
    else:
        return "较低"


def style_excel(filename):
    """
    美化Excel表格
    """
    print(f"\n{'=' * 80}")
    print("正在美化Excel表格...")
    print(f"{'=' * 80}\n")

    wb = load_workbook(filename)
    ws = wb.active

    # 定义样式
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

    category_fills = {
        "金属硫化物": PatternFill(
            start_color="FFE699", end_color="FFE699", fill_type="solid"
        ),
        "金属氧化物": PatternFill(
            start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"
        ),
        "金属硫氧化物": PatternFill(
            start_color="F4B084", end_color="F4B084", fill_type="solid"
        ),
        "氮化物": PatternFill(
            start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"
        ),
        "碳化物": PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        ),
        "硒化物": PatternFill(
            start_color="FFD966", end_color="FFD966", fill_type="solid"
        ),
        "碲化物": PatternFill(
            start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"
        ),
        "卤化物": PatternFill(
            start_color="C9DAF8", end_color="C9DAF8", fill_type="solid"
        ),
        "磷化物": PatternFill(
            start_color="D5A6BD", end_color="D5A6BD", fill_type="solid"
        ),
        "砷化物": PatternFill(
            start_color="B6D7A8", end_color="B6D7A8", fill_type="solid"
        ),
    }

    data_font = Font(name="微软雅黑", size=10)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # 设置表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border

    # 设置数据行样式
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        category = ws.cell(row=row_idx, column=1).value

        for col_idx, cell in enumerate(row, start=1):
            cell.font = data_font
            cell.border = thin_border

            # 根据分类设置背景色
            if category in category_fills:
                cell.fill = category_fills[category]

            # 设置对齐方式
            if col_idx in [1, 2, 3, 4, 5, 9, 16, 17, 21, 22]:  # 文本列左对齐
                cell.alignment = alignment_left
            else:
                cell.alignment = alignment_center

    # 调整列宽
    column_widths = {
        "A": 15,  # 分类
        "B": 15,  # 化学系统
        "C": 15,  # 材料ID
        "D": 18,  # 化学式
        "E": 20,  # 元素组成
        "F": 10,  # 元素数
        "G": 10,  # 原子数
        "H": 12,  # 带隙
        "I": 12,  # 直接带隙
        "J": 15,  # CBM
        "K": 15,  # VBM
        "L": 15,  # 费米能级
        "M": 18,  # 形成能
        "N": 20,  # 能量高于凸包
        "O": 12,  # 晶系
        "P": 15,  # 空间群
        "Q": 14,  # 密度
        "R": 14,  # 体积
        "S": 16,  # 光电应用潜力
        "T": 16,  # 光催化应用潜力
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置行高
    ws.row_dimensions[1].height = 30  # 表头行高
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 25  # 数据行高

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存
    wb.save(filename)
    print("✓ Excel美化完成")


def save_summary_report(df, filename="主流半导体材料数据摘要.txt"):
    """
    生成摘要报告
    """
    print(f"\n{'=' * 80}")
    print("正在生成数据摘要报告...")
    print(f"{'=' * 80}\n")

    report = []
    report.append("=" * 80)
    report.append("Materials Project - 主流半导体材料数据摘要报告")
    report.append("=" * 80)
    report.append(f"生成时间: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}")
    report.append("数据来源: Materials Project API")
    report.append("作者: Luffy.Solution")
    report.append("=" * 80)
    report.append("")

    # 总体统计
    report.append("📊 总体统计")
    report.append("-" * 80)
    report.append(f"材料总数: {len(df)}")
    report.append(f"材料类别: {df['分类'].nunique()}")
    report.append("")

    # 分类统计
    report.append("📋 按分类统计")
    report.append("-" * 80)
    category_counts = df["分类"].value_counts()
    for category, count in category_counts.items():
        report.append(f"  {category:<15} {count:>3} 个材料")
    report.append("")

    # 带隙分布
    report.append("⚡ 带隙分布统计")
    report.append("-" * 80)
    bg_data = df[df["带隙 (eV)"] != "N/A"]["带隙 (eV)"].astype(float)
    if len(bg_data) > 0:
        report.append(f"  平均带隙: {bg_data.mean():.3f} eV")
        report.append(f"  最小带隙: {bg_data.min():.3f} eV")
        report.append(f"  最大带隙: {bg_data.max():.3f} eV")
        report.append(f"  中位带隙: {bg_data.median():.3f} eV")

        # 带隙区间分布
        report.append("")
        report.append("  带隙区间分布:")
        report.append(
            f"    0.0-1.0 eV (红外):        {len(bg_data[bg_data < 1.0]):>3} 个"
        )
        report.append(
            f"    1.0-2.0 eV (近红外-可见): {len(bg_data[(bg_data >= 1.0) & (bg_data < 2.0)]):>3} 个"
        )
        report.append(
            f"    2.0-3.0 eV (可见-紫外):   {len(bg_data[(bg_data >= 2.0) & (bg_data < 3.0)]):>3} 个"
        )
        report.append(
            f"    3.0+ eV (紫外):          {len(bg_data[bg_data >= 3.0]):>3} 个"
        )
    report.append("")

    # 直接带隙统计
    report.append("🔬 带隙类型统计")
    report.append("-" * 80)
    direct_counts = df["直接带隙"].value_counts()
    for gap_type, count in direct_counts.items():
        report.append(f"  {gap_type}直接带隙: {count} 个材料")
    report.append("")

    # 晶系分布
    report.append("🔷 晶系分布")
    report.append("-" * 80)
    crystal_counts = df["晶系"].value_counts()
    for crystal, count in crystal_counts.items():
        report.append(f"  {crystal:<15} {count:>3} 个材料")
    report.append("")

    # 应用潜力统计
    report.append("🌟 光电应用潜力统计")
    report.append("-" * 80)
    pv_counts = df["光电应用潜力"].value_counts()
    for potential, count in pv_counts.items():
        report.append(f"  {potential:<10} {count:>3} 个材料")
    report.append("")

    report.append("💧 光催化应用潜力统计")
    report.append("-" * 80)
    pc_counts = df["光催化应用潜力"].value_counts()
    for potential, count in pc_counts.items():
        report.append(f"  {potential:<10} {count:>3} 个材料")
    report.append("")

    # TOP材料推荐
    report.append("⭐ TOP 10 光伏候选材料（按带隙排序）")
    report.append("-" * 80)
    pv_materials = df[df["光电应用潜力"].isin(["优秀", "良好"])]
    if len(pv_materials) > 0:
        pv_materials_sorted = pv_materials.nsmallest(10, "带隙 (eV)")
        for idx, (_, row) in enumerate(pv_materials_sorted.iterrows(), 1):
            report.append(
                f"  {idx:>2}. {row['化学式']:<15} 带隙: {row['带隙 (eV)']} eV ({row['分类']})"
            )
    report.append("")

    report.append("🌊 TOP 10 光催化候选材料")
    report.append("-" * 80)
    pc_materials = df[df["光催化应用潜力"].isin(["优秀", "良好"])]
    if len(pc_materials) > 0:
        pc_materials_sorted = pc_materials.head(10)
        for idx, (_, row) in enumerate(pc_materials_sorted.iterrows(), 1):
            report.append(
                f"  {idx:>2}. {row['化学式']:<15} 带隙: {row['带隙 (eV)']} eV ({row['分类']})"
            )
    report.append("")

    report.append("=" * 80)
    report.append("报告结束")
    report.append("=" * 80)

    # 保存报告
    with open(filename, "w", encoding="utf-8") as f:
        f.write("\n".join(report))

    print(f"✓ 摘要报告已保存: {filename}")

    # 同时打印到控制台
    print("\n" + "\n".join(report))


# ============================================================================
# 主程序
# ============================================================================


def main():
    """主函数"""
    start_time = time.time()

    print("开始时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print()

    # 第一步：搜索各类半导体材料
    all_materials = []

    for category_name, category_info in SEMICONDUCTOR_CATEGORIES.items():
        materials = search_semiconductors_by_category(
            category_name,
            category_info["elements"],
            limit_per_system=5,  # 每个化学系统获取5个材料
        )
        all_materials.extend(materials)

    print(f"\n{'=' * 80}")
    print(f"第一阶段完成：共搜索到 {len(all_materials)} 个半导体材料")
    print(f"{'=' * 80}")

    if len(all_materials) == 0:
        print("\n❌ 未找到符合条件的材料，程序退出")
        return

    # 第二步：获取详细电子结构信息
    enriched_materials = enrich_material_data(all_materials[:120])  # 限制120个以免超时

    # 第三步：创建DataFrame
    df = create_dataframe(enriched_materials)

    # 第四步：保存数据
    excel_file = "主流半导体材料数据库.xlsx"
    json_file = "主流半导体材料数据库.json"

    print(f"\n{'=' * 80}")
    print("正在保存数据...")
    print(f"{'=' * 80}\n")

    # 保存Excel
    df.to_excel(excel_file, index=False, engine="openpyxl")
    print(f"✓ Excel数据已保存: {excel_file}")

    # 美化Excel
    style_excel(excel_file)

    # 保存JSON
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(enriched_materials, f, indent=2, ensure_ascii=False)
    print(f"✓ JSON数据已保存: {json_file}")

    # 第五步：生成摘要报告
    save_summary_report(df)

    # 完成
    elapsed_time = time.time() - start_time

    print(f"\n{'=' * 80}")
    print("✅ 所有任务完成！")
    print(f"{'=' * 80}")
    print(f"总材料数: {len(df)}")
    print(f"总耗时: {elapsed_time:.1f} 秒 ({elapsed_time / 60:.1f} 分钟)")
    print(f"Excel文件: {excel_file}")
    print(f"JSON文件: {json_file}")
    print("摘要报告: 主流半导体材料数据摘要.txt")
    print(f"{'=' * 80}")


if __name__ == "__main__":
    main()
